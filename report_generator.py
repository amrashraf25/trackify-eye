"""
report_generator.py
-------------------
Generates a single styled Excel report containing ALL rows from the data.
"""

import logging
from pathlib import Path
from typing import Any, Dict, List

import openpyxl
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── Style constants ───────────────────────────────────────────────────────────
HEADER_BG  = "2E4057"   # dark blue-grey
HEADER_FG  = "FFFFFF"   # white
ROW_ALT_BG = "EAF0FB"   # light blue (alternate rows)
BORDER_CLR = "BDC3C7"

_thin = Side(style="thin", color=BORDER_CLR)
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def generate_report(
    rows: List[Dict[str, Any]],
    output_path: str,
    title: str = "Report",
    template_path=None,          # kept for API compatibility, not used for xlsx
) -> None:
    """
    Build a styled Excel (.xlsx) report from *rows* and save to *output_path*.
    """
    if not rows:
        raise ValueError("No rows to include in report.")

    headers = list(rows[0].keys())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = title[:31]   # sheet name max 31 chars

    # ── Title row ─────────────────────────────────────────────────────────────
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=len(headers))
    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font      = Font(name="Calibri", size=14, bold=True, color="2E4057")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    # ── Header row ────────────────────────────────────────────────────────────
    for col_i, header in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_i, value=header)
        cell.font      = Font(name="Calibri", size=10, bold=True, color=HEADER_FG)
        cell.fill      = PatternFill("solid", fgColor=HEADER_BG)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = _border
    ws.row_dimensions[2].height = 20

    # ── Data rows ─────────────────────────────────────────────────────────────
    for row_i, row in enumerate(rows, start=3):
        is_alt = (row_i % 2 == 0)
        for col_i, header in enumerate(headers, start=1):
            value = row.get(header, "")
            cell  = ws.cell(row=row_i, column=col_i, value=value)
            cell.font      = Font(name="Calibri", size=10)
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border    = _border
            if is_alt:
                cell.fill = PatternFill("solid", fgColor=ROW_ALT_BG)
        ws.row_dimensions[row_i].height = 16

    # ── Auto column widths ────────────────────────────────────────────────────
    for col_i, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_i)
        # Max content width in this column
        max_len = len(str(header))
        for row in rows:
            val = str(row.get(header, ""))
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # ── Freeze header rows ────────────────────────────────────────────────────
    ws.freeze_panes = "A3"

    # ── Auto-filter on header row ─────────────────────────────────────────────
    ws.auto_filter.ref = (
        f"A2:{get_column_letter(len(headers))}{len(rows) + 2}"
    )

    # ── Save ──────────────────────────────────────────────────────────────────
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    logger.info("Report saved: %s  (%d rows, %d columns)",
                output_path, len(rows), len(headers))
